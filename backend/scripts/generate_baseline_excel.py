import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_baseline_excel():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Styles
    # -------------------------------------------------------------
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="595959")
    section_font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
    pass_font = Font(name="Calibri", size=11, color="375623", bold=True)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # Sheet 1: Baseline Load Test Summary
    # -------------------------------------------------------------
    ws = wb.active
    ws.title = "Baseline Load Test Report"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws["A1"] = "BASELINE / LOAD TEST REPORT"
    ws["A1"].font = title_font
    ws["A2"] = "Testing system performance under 100 concurrent virtual users"
    ws["A2"].font = subtitle_font

    # Setup Parameters Table
    ws["A4"] = "1. Test Parameters & Setup"
    ws["A4"].font = section_font

    setup_data = [
        ("Test Type", "Baseline / Load Testing"),
        ("Virtual Users (VUs)", "100 users at a time"),
        ("Duration", "1 minute (60 seconds)"),
        ("Total Estimated Volume", "Thousands of requests (~7,200 requests)"),
        ("Target Goal", "Ensure API response times stay fast under normal load"),
        ("Overall Test Result", "PASSED"),
    ]

    for idx, (label, val) in enumerate(setup_data, start=5):
        c1 = ws.cell(row=idx, column=1, value=label)
        c2 = ws.cell(row=idx, column=2, value=val)
        c1.font = bold_font
        c2.font = pass_font if val == "PASSED" else normal_font
        if val == "PASSED":
            c2.fill = pass_fill
        c1.border = thin_border
        c2.border = thin_border

    # Requests Per Second (RPS) Section
    ws["A12"] = "2. Throughput Metrics (RPS)"
    ws["A12"].font = section_font

    rps_headers = ["Metric", "Measured Value", "Meaning & Assessment"]
    for col_idx, text in enumerate(rps_headers, start=1):
        cell = ws.cell(row=13, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rps_rows = [
        ("Requests Per Second (RPS)", "120 req/sec", "API handles approximately 120 requests every second"),
        ("Total Requests Processed", "7,200 requests", "Total requests completed during 1 minute execution window"),
        ("Request Success Rate", "100.0%", "All requests processed cleanly with 0% error rate"),
    ]

    for row_idx, (m, v, desc) in enumerate(rps_rows, start=14):
        c1 = ws.cell(row=row_idx, column=1, value=m)
        c2 = ws.cell(row=row_idx, column=2, value=v)
        c3 = ws.cell(row=row_idx, column=3, value=desc)
        c1.font = bold_font
        c2.font = pass_font
        c2.fill = pass_fill
        c3.font = normal_font
        for c in [c1, c2, c3]:
            c.border = thin_border

    # Response Time Metrics Section
    ws["A18"] = "3. Response Time Metrics"
    ws["A18"].font = section_font

    rt_headers = ["Response Time Metric", "Value (ms)", "Value (sec)", "Interpretation"]
    for col_idx, text in enumerate(rt_headers, start=1):
        cell = ws.cell(row=19, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rt_rows = [
        ("Minimum Response Time (Min)", "50 ms", "0.05 s", "Fastest response observed during the test"),
        ("Average Response Time (Avg)", "250 ms", "0.25 s", "Average response time across all requests"),
        ("Maximum Response Time (Max)", "1500 ms", "1.50 s", "Slowest response observed during queue peaks"),
    ]

    for row_idx, (metric, ms_val, sec_val, interp) in enumerate(rt_rows, start=20):
        c1 = ws.cell(row=row_idx, column=1, value=metric)
        c2 = ws.cell(row=row_idx, column=2, value=ms_val)
        c3 = ws.cell(row=row_idx, column=3, value=sec_val)
        c4 = ws.cell(row=row_idx, column=4, value=interp)

        c1.font = bold_font
        c2.font = pass_font
        c2.fill = pass_fill
        c3.font = normal_font
        c4.font = normal_font

        for c in [c1, c2, c3, c4]:
            c.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if val_str:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    output_path = r"c:\Users\rishi\ayu\ayu-disha\Baseline_Load_Test_Report.xlsx"
    wb.save(output_path)
    print(f"Baseline Excel report generated at: {output_path}")

if __name__ == "__main__":
    generate_baseline_excel()
