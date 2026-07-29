import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_excel_report():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Styles Definition
    # -------------------------------------------------------------
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Deep Navy
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="595959")
    section_font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    success_font = Font(name="Calibri", size=11, color="375623", bold=True)
    
    warning_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    warning_font = Font(name="Calibri", size=11, color="C65911", bold=True)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # Sheet 1: Executive Summary
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True

    ws1["A1"] = "AYU DISHA - API BASELINE LOAD TEST REPORT"
    ws1["A1"].font = title_font
    ws1["A2"] = "Automated Performance & Capacity Test Execution Summary"
    ws1["A2"].font = subtitle_font

    ws1["A4"] = "Test Environment & Parameters"
    ws1["A4"].font = section_font

    params = [
        ("Target Endpoint URL", "https://ayu-disha.onrender.com/"),
        ("Test Date", "2026-07-29"),
        ("Concurrent Virtual Users (VUs)", 100),
        ("Target Test Duration", "60 seconds"),
        ("Actual Test Duration", "62.36 seconds"),
        ("Overall Test Result", "PASSED (100% Success Rate)"),
    ]

    for idx, (label, val) in enumerate(params, start=5):
        c1 = ws1.cell(row=idx, column=1, value=label)
        c2 = ws1.cell(row=idx, column=2, value=val)
        c1.font = bold_font
        c2.font = success_font if "PASSED" in str(val) else normal_font
        c1.border = thin_border
        c2.border = thin_border

    ws1["A12"] = "Key Performance Indicators (KPIs)"
    ws1["A12"].font = section_font

    headers_kpi = ["Metric Name", "Value", "Unit", "Status / Assessment"]
    for col_idx, text in enumerate(headers_kpi, start=1):
        cell = ws1.cell(row=13, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("Total Requests Processed", 2191, "requests", "Completed Load Cycle", success_fill, success_font),
        ("Requests Per Second (RPS)", 35.13, "req/sec", "Baseline Throughput", success_fill, success_font),
        ("Success Rate (HTTP 200)", "100.0%", "percentage", "2,191 of 2,191 requests passed", success_fill, success_font),
        ("Error / Timeout Rate", "0.0%", "percentage", "Zero errors or timeouts", success_fill, success_font),
        ("Average Response Time", 2782.90, "ms", "Average server latency under 100 VUs", normal_font, normal_font),
        ("Minimum Response Time", 153.31, "ms", "Fastest response observed", success_fill, success_font),
        ("P95 Latency", 7487.63, "ms", "95% of requests completed < 7.48s", success_fill, success_font),
        ("P99 Latency", 8943.01, "ms", "99% of requests completed < 8.94s", success_fill, success_font),
        ("Maximum Response Time", 15737.75, "ms", "Peak latency under max queue load", normal_font, normal_font),
    ]

    for row_idx, (name, val, unit, assessment, fill_style, font_style) in enumerate(kpis, start=14):
        c1 = ws1.cell(row=row_idx, column=1, value=name)
        c2 = ws1.cell(row=row_idx, column=2, value=val)
        c3 = ws1.cell(row=row_idx, column=3, value=unit)
        c4 = ws1.cell(row=row_idx, column=4, value=assessment)
        
        c1.font = bold_font
        c2.font = font_style if font_style != normal_font else normal_font
        c3.font = normal_font
        c4.font = font_style if font_style != normal_font else normal_font

        if fill_style != normal_font:
            c2.fill = fill_style
            c4.fill = fill_style

        for c in [c1, c2, c3, c4]:
            c.border = thin_border

    # -------------------------------------------------------------
    # Sheet 2: Detailed Latency Breakdown
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Latency Breakdown")
    ws2.views.sheetView[0].showGridLines = True

    ws2["A1"] = "RESPONSE TIME & LATENCY DISTRIBUTION"
    ws2["A1"].font = title_font

    headers_lat = ["Latency Metric", "Value (ms)", "Value (sec)", "Description"]
    for col_idx, text in enumerate(headers_lat, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    latency_data = [
        ("Minimum Latency", 153.31, 0.15, "Best case execution under light queueing"),
        ("Average Latency", 2782.90, 2.78, "Average execution time across 2,191 requests"),
        ("P95 Latency", 7487.63, 7.49, "95% of all incoming requests completed within 7.49 seconds"),
        ("P99 Latency", 8943.01, 8.94, "99% of all incoming requests completed within 8.94 seconds"),
        ("Maximum Latency", 15737.75, 15.74, "Worst case execution during peak connection queueing"),
    ]

    for row_idx, (metric, ms_val, sec_val, desc) in enumerate(latency_data, start=4):
        c1 = ws2.cell(row=row_idx, column=1, value=metric)
        c2 = ws2.cell(row=row_idx, column=2, value=ms_val)
        c3 = ws2.cell(row=row_idx, column=3, value=sec_val)
        c4 = ws2.cell(row=row_idx, column=4, value=desc)

        c1.font = bold_font
        c2.font = normal_font
        c3.font = normal_font
        c4.font = normal_font

        for c in [c1, c2, c3, c4]:
            c.border = thin_border

    # -------------------------------------------------------------
    # Sheet 3: Response Status Code Analysis
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Status Breakdown")
    ws3.views.sheetView[0].showGridLines = True

    ws3["A1"] = "HTTP STATUS CODE BREAKDOWN"
    ws3["A1"].font = title_font

    headers_status = ["HTTP Result Code", "Count", "Percentage", "Category", "Impact & Assessment"]
    for col_idx, text in enumerate(headers_status, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_data = [
        ("HTTP 200 OK", 2191, "100.0%", "Success", "100% of all requests successfully processed and returned status 200", success_fill, success_font),
        ("Errors / Timeouts", 0, "0.0%", "None", "Zero connection drops or timeouts occurred", success_fill, success_font),
        ("Total Requests", 2191, "100.0%", "Total Volume", "Sum of all attempted HTTP connections", header_fill, header_font),
    ]

    for row_idx, (code, count, pct, cat, impact, fill_s, font_s) in enumerate(status_data, start=4):
        c1 = ws3.cell(row=row_idx, column=1, value=code)
        c2 = ws3.cell(row=row_idx, column=2, value=count)
        c3 = ws3.cell(row=row_idx, column=3, value=pct)
        c4 = ws3.cell(row=row_idx, column=4, value=cat)
        c5 = ws3.cell(row=row_idx, column=5, value=impact)

        c1.font = bold_font
        c2.font = font_s if font_s != normal_font else normal_font
        c3.font = font_s if font_s != normal_font else normal_font
        c4.font = font_s if font_s != normal_font else normal_font
        c5.font = font_s if font_s != normal_font else normal_font

        if fill_s != normal_font:
            for c in [c1, c2, c3, c4, c5]:
                c.fill = fill_s
                c.font = font_s

        for c in [c1, c2, c3, c4, c5]:
            c.border = thin_border

    # -------------------------------------------------------------
    # Sheet 4: Recommendations & Optimization Plan
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Recommendations")
    ws4.views.sheetView[0].showGridLines = True

    ws4["A1"] = "RECOMMENDATIONS FOR CONCURRENCY & LATENCY OPTIMIZATION"
    ws4["A1"].font = title_font

    headers_rec = ["Priority", "Area", "Recommended Action", "Expected Impact"]
    for col_idx, text in enumerate(headers_rec, start=1):
        cell = ws4.cell(row=3, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    recs = [
        ("HIGH", "Uvicorn Worker Concurrency", "Configure Gunicorn/Uvicorn to spawn 4-8 worker processes: `gunicorn main:app -w 4 -k uvicorn.workers.UvicornWorker`", "Eliminates worker queue timeouts and increases RPS by 3x-4x"),
        ("HIGH", "Cloud Instance Scaling", "Upgrade Render compute tier or configure auto-scaling for peak concurrency", "Prevents CPU throttling under 100+ concurrent users"),
        ("MEDIUM", "Database Connection Pool", "Configure Motor / PyMongo maxPoolSize=100 in database.py", "Prevents database lock waiting during high concurrent DB writes/reads"),
        ("LOW", "API Caching Strategy", "Implement Redis or in-memory caching for static/read-heavy endpoints", "Reduces response times to < 50ms for cached routes"),
    ]

    for row_idx, (prio, area, action, impact) in enumerate(recs, start=4):
        c1 = ws4.cell(row=row_idx, column=1, value=prio)
        c2 = ws4.cell(row=row_idx, column=2, value=area)
        c3 = ws4.cell(row=row_idx, column=3, value=action)
        c4 = ws4.cell(row=row_idx, column=4, value=impact)

        c1.font = Font(name="Calibri", size=11, bold=True, color="C65911" if prio=="HIGH" else "1F4E78")
        c2.font = bold_font
        c3.font = normal_font
        c4.font = normal_font

        for c in [c1, c2, c3, c4]:
            c.border = thin_border

    # Auto-adjust column widths across all sheets
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if val_str:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to workspace root
    output_path = r"c:\Users\rishi\ayu\ayu-disha\Load_Test_Report_AyuDisha_Passed.xlsx"
    wb.save(output_path)
    print(f"Excel report successfully generated at: {output_path}")

if __name__ == "__main__":
    create_excel_report()
